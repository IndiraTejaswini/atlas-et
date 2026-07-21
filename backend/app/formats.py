"""Multi-format document readers.

Turns heterogeneous industrial file formats — PDF drawings/reports, CSV logs,
Excel registers, .eml email archives, and plain text/markdown — into a single
normalized (frontmatter, body) representation so the rest of the pipeline
(extraction, graph, retrieval) is format-agnostic.

Everything downstream sees the same Document shape regardless of source format.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from email import message_from_bytes
from email.policy import default as email_default

logger = logging.getLogger("atlas.formats")

SUPPORTED = {
    ".md": "markdown", ".txt": "text", ".pdf": "pdf",
    ".csv": "csv", ".xlsx": "xlsx", ".xls": "xlsx", ".eml": "email",
    ".png": "image", ".jpg": "image", ".jpeg": "image", ".tif": "image", ".tiff": "image",
}


def detect_format(filename: str) -> str:
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return SUPPORTED.get(ext, "text")


def _clean(s: str) -> str:
    return re.sub(r"[ \t]+", " ", s).strip()


MAX_PDF_PAGES = 500  # bounds worst-case CPU/memory from a maliciously huge page count


def read_pdf(data: bytes) -> tuple[dict, str]:
    from pypdf import PdfReader

    from . import ocr

    reader = PdfReader(io.BytesIO(data))
    meta = {"type": "drawing" if _looks_like_drawing(reader) else "inspection"}
    info = reader.metadata or {}
    if info.get("/Title"):
        meta["title"] = str(info["/Title"])
    total_pages = len(reader.pages)
    pages = []
    for i, page in enumerate(reader.pages[:MAX_PDF_PAGES]):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"## Page {i + 1}\n\n{_clean(text)}")
    if total_pages > MAX_PDF_PAGES:
        pages.append(f"*(Truncated: {total_pages} pages in source, only the first {MAX_PDF_PAGES} were processed.)*")
    body = "\n\n".join(pages)
    if not body.strip():  # image-only / scanned PDF → OCR fallback
        body = ocr.ocr_pdf_pages(data) or (
            "*(Scanned image PDF with no embedded text — OCR engine not installed in this environment.)*")
    return meta, body


def read_image(data: bytes) -> tuple[dict, str]:
    from . import ocr

    text = ocr.ocr_image_bytes(data)
    body = f"## Scanned form (OCR)\n\n{text}" if text.strip() else (
        "*(Image received — OCR engine not installed in this environment.)*")
    return {"type": "inspection"}, body


def _looks_like_drawing(reader) -> bool:
    head = (reader.pages[0].extract_text() or "").lower() if reader.pages else ""
    return any(k in head for k in ("p&id", "drawing", "isometric", "equipment register"))


def read_csv(data: bytes) -> tuple[dict, str]:
    text = data.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return {"type": "datasheet"}, _table_to_markdown(rows)


def read_xlsx(data: bytes) -> tuple[dict, str]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(cell.strip() for cell in r)]
        if not rows:
            continue
        parts.append(f"## Sheet: {ws.title}\n\n{_table_to_markdown(rows)}")
    wb.close()
    return {"type": "datasheet"}, "\n\n".join(parts) or "*(empty workbook)*"


def read_email(data: bytes) -> tuple[dict, str]:
    msg = message_from_bytes(data, policy=email_default)
    meta = {"type": "email"}
    if msg["subject"]:
        meta["title"] = str(msg["subject"])
    if msg["date"]:
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(msg["date"]))
        if m:
            meta["date"] = m.group(0)
    header = []
    for h in ("From", "To", "Date", "Subject"):
        if msg[h]:
            header.append(f"**{h}:** {msg[h]}")
    try:
        body_part = msg.get_body(preferencelist=("plain",))
        body = body_part.get_content() if body_part else ""
    except Exception:
        logger.warning("could not extract .eml body — ingesting headers only", exc_info=True)
        body = ""
    return meta, "\n".join(header) + "\n\n" + _clean(body)


def _table_to_markdown(rows: list[list[str]], max_rows: int = 200) -> str:
    rows = [r for r in rows if r][:max_rows]
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = rows[0]
    out = ["| " + " | ".join(header) + " |", "|" + "|".join(["---"] * width) + "|"]
    for r in rows[1:]:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def read_any(filename: str, data: bytes) -> tuple[dict, str, str]:
    """Return (extra_meta, body, format_label). Raw text/markdown handled by caller."""
    fmt = detect_format(filename)
    if fmt == "pdf":
        meta, body = read_pdf(data)
    elif fmt == "csv":
        meta, body = read_csv(data)
    elif fmt == "xlsx":
        meta, body = read_xlsx(data)
    elif fmt == "email":
        meta, body = read_email(data)
    elif fmt == "image":
        meta, body = read_image(data)
    else:
        return {}, data.decode("utf-8", errors="replace"), fmt
    return meta, body, fmt
